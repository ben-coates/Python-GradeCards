from openpyxl.styles import Font, Alignment, Border, Side

def infoRow(sheet, row, label, info):
    sheet["A" + str(row)] = label
    sheet["A" + str(row)].alignment = Alignment(horizontal="right", vertical="center")
    sheet["A" + str(row)].font = Font(bold=True)
    sheet["B" + str(row)] = info
    sheet["B" + str(row)].alignment = Alignment(horizontal="left", vertical="center")


def classHeaderCell(sheet, column, row, label):
    sheet[column + str(row)] = label
    sheet[column + str(row)].font = Font(bold=True)
    sheet[column + str(row)].alignment = Alignment(horizontal="center", vertical="center")

def formatGradecard(sheet, student, strw):
    couponsList = [ chr(9312), 
                    chr(9312) + " " + chr(9313),
                    chr(9312) + " " + chr(9313) + " " + chr(9314),
                    chr(9312) + " " + chr(9313) + " " + chr(9314) + " " + chr(9315),
                    chr(9312) + " " + chr(9313) + " " + chr(9314) + " " + chr(9315) + " " + chr(9316),
                    "None"]

    # Student Info
    infoRow(sheet, strw, "ID:", student.id)
    infoRow(sheet, strw + 1, "Date:", student.date)
    infoRow(sheet, strw + 2, "Name:", student.name)
    infoRow(sheet, strw + 3, "Grade:", student.gradelevel)
    infoRow(sheet, strw + 6, "GPA:", student.gpa)
    infoRow(sheet, strw + 7, "Effort", student.effort)
    infoRow(sheet, strw + 8, "Status:", student.eligibility)

    # Class Info 
    classHeaderCell(sheet, "D", strw, "Class")
    classHeaderCell(sheet, "E", strw, "Grade")
    classHeaderCell(sheet, "G", strw, "Effort")
    classHeaderCell(sheet, "H", strw, "Coupons")
    sheet.merge_cells("E" + str(strw) + ":F" + str(strw))
    sheet["H" + str(strw + 1)] = couponsList[student.coupons - 1]
    sheet.merge_cells("H" + str(strw + 1) + ":H" + str(strw + 8))
    sheet["H" + str(strw + 1)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90)
    sheet["H" + str(strw + 1)].font = Font(size=16)

    for index, i in enumerate(student.classes):
        row = strw + index + 1
        sheet["D" + str(row)] = i.get("classname")
        sheet["E" + str(row)] = i.get("lettergrade")
        sheet["E" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
        sheet["F" + str(row)] = i.get("percentage")
        sheet["F" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
        sheet["G" + str(row)] = i.get("effort")
        sheet["G" + str(row)].alignment = Alignment(horizontal="center", vertical="center")

    # Create Borders
    # This could be done better. It seems rathe clunky, but currently gets the job done.
    mediumSide = Side(color="000000", style="medium")
    thinSide =  Side(color="000000", style="thin")

    for i in range(9):

        if i == 0:
            sheet["A" + str(strw + i)].border = Border(left=mediumSide, top=mediumSide)
            sheet["B" + str(strw + i)].border = Border(right=mediumSide, top=mediumSide)
            sheet["C" + str(strw + i)].border = Border(top=mediumSide)
            sheet["D" + str(strw + i)].border = Border(left=mediumSide, top=mediumSide, bottom=thinSide)
            sheet["E" + str(strw + i)].border = Border(top=mediumSide, bottom=thinSide)
            sheet["F" + str(strw + i)].border = Border(top=mediumSide, bottom=thinSide)
            sheet["G" + str(strw + i)].border = Border(right=mediumSide, top=mediumSide, bottom=thinSide)
            sheet["H" + str(strw + i)].border = Border(right=mediumSide, top=mediumSide, bottom=thinSide)
        elif i == 8:
            sheet["A" + str(strw + i)].border = Border(left=mediumSide, bottom=mediumSide)
            sheet["B" + str(strw + i)].border = Border(right=mediumSide, bottom=mediumSide)
            sheet["C" + str(strw + i)].border = Border(bottom=mediumSide)
            sheet["D" + str(strw + i)].border = Border(left=mediumSide, bottom=mediumSide)
            sheet["E" + str(strw + i)].border = Border(bottom=mediumSide)
            sheet["F" + str(strw + i)].border = Border(bottom=mediumSide)
            sheet["G" + str(strw + i)].border = Border(right=mediumSide, bottom=mediumSide)
            sheet["H" + str(strw + i)].border = Border(right=mediumSide, bottom=mediumSide)
        else:
            sheet["A" + str(strw + i)].border = Border(left=mediumSide)
            sheet["B" + str(strw + i)].border = Border(right=mediumSide)
            sheet["D" + str(strw + i)].border = Border(left=mediumSide)
            sheet["G" + str(strw + i)].border = Border(right=mediumSide)
            sheet["H" + str(strw + i)].border = Border(right=mediumSide)

def runGradeCards(students, workbook):
    sheet = workbook.create_sheet(title="Gradecards")

    startrow = 1
    for student in students:
        formatGradecard(sheet, student, startrow)
        startrow = startrow + 11

    sheet.column_dimensions["A"].width = 7
    sheet.column_dimensions["B"].width = 25
    sheet.column_dimensions["C"].width = 1
    sheet.column_dimensions["D"].width = 25
    sheet.column_dimensions["E"].width = 5
    sheet.column_dimensions["F"].width = 5
    sheet.column_dimensions["G"].width = 7
    sheet.column_dimensions["H"].width = 10
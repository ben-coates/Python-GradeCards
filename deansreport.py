# deansreport.py
# Functions for exporting data to excel for the deans report

# Import openpyxl styles formatting
from openpyxl import Workbook
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# Function Name: headerCell()
# Description: Function for creating a header cell in the deans report
# Parameters:
#   sheet   openpyxl sheet object
#   cell    openpyxl cell object
#   text    string
#   width   int
#   color   hex color value 
def headerCell(sheet: object, cell: object, text: str, width: int, color: str):
    sheet[cell] = text
    sheet[cell].alignment = Alignment(horizontal='center', vertical='center')
    sheet[cell].fill = PatternFill("solid", start_color=color)
    sheet[cell].font = Font(color="FFFFFF")
    sheet.column_dimensions[cell[0]].width = width

# Function Name: bodyRow()
# Description: Function for creating a body cell in the deans report
# Parameters:
#   sheet   openpyxl sheet object
#   cell    openpyxl cell object
#   text    string
#   width   int
#   color   hex color value 
def bodyRow(sheet: object, row: int, student: object):
    sheet.row_dimensions[row].height = 12

    row = str(row)
    sheet["A" + row] = student.name
    sheet["B" + row] = student.gpa
    sheet["C" + row] = student.eligibility

    sheet["B" + row].alignment = Alignment(horizontal='center')
    sheet["C" + row].alignment = Alignment(horizontal='center')

    for cell in sheet[row]:
        cell.fill = PatternFill("solid", start_color="EFEFEF")

# Function Name: runDeansReport
# Desctiption: Function for creating the deans report in the excel file
# Parameters:
#   report      Python List of student objects
#   workbook    openpyxl workbook object
def runDeansReport(report: list, workbook: object):
    # Set the sort order of eligibility
    eligibility_order = ['Honor Roll', 'Eligible', 'Limited', 'Ineligible']
    
    # Sort methods starting with the third level report and ending with the primary sort method
    # Sorted primarily on eligibility (eligiblity_order), next on gpa (highest to lowest) and last on name (alphabetical)
    report = sorted(report, key=lambda s: s.name)
    report = sorted(report, key=lambda s: s.gpa, reverse=True)
    report = sorted(report, key=lambda s: eligibility_order.index(s.eligibility))

    # Create openpyxl Worksheet object using openpyxl Workbook object. Titled Deans Report
    sheet = workbook.active
    sheet.title = "Deans Report"

    # Create and format 3 header using the headerCell function. Student Name, GPA and Status (Eligibility)
    headerCell(sheet, "A1", "Student Name", 40, "008000")
    headerCell(sheet, "B1", "GPA", 8, "008000")
    headerCell(sheet, "C1", "Status", 15, "008000")

    # Add Report Data to Worksheet
    for index, student in enumerate(report):
        bodyRow(sheet, index + 2, student)
    
    sheet.print_area = 'A1:C' + str(len(report) + 1)

